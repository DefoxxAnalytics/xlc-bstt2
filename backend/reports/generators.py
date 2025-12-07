"""
Report generators for BSTT Compliance Dashboard.
Generates Excel reports matching BSTT-rpt.xlsx format.
"""
from io import BytesIO
from datetime import datetime
import pandas as pd
from django.db.models import QuerySet, Sum, Count, Q, Min, Max
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.models import TimeEntry
from kpis.calculator import KPICalculator


class BSTTReportGenerator:
    """Generate Excel reports for BSTT compliance data matching BSTT-rpt.xlsx format."""

    # Enrollment threshold - employees with more provisional entries need fingerprint enrollment
    ENROLLMENT_THRESHOLD = 2

    def __init__(self, queryset: QuerySet, year: int = None, filters: dict = None):
        self.qs = queryset
        self.year = year or datetime.now().year
        self.filters = filters or {}

    def generate_full_report(self) -> BytesIO:
        """Generate full BSTT Excel report with all sheets matching BSTT-rpt.xlsx."""
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Metadata sheets
            self._write_checking_file(writer)
            self._write_directions(writer)

            # Core data sheet
            self._write_data(writer)

            # Entry type analysis sheets
            self._write_all_pivot(writer)
            self._write_prov_weekly(writer)
            self._write_write_ins(writer)
            self._write_provisional_detail(writer)

            # Office-specific sheets
            offices = self._get_offices()
            for office in offices:
                self._write_office_sheet(writer, office)

            # Apply formatting to all sheets
            for sheet_name in writer.sheets:
                self._apply_formatting(writer.sheets[sheet_name])

        output.seek(0)
        return output

    # ============================================
    # Helper Methods
    # ============================================

    def _get_unique_weeks(self):
        """Get sorted list of unique weeks in the data."""
        weeks = self.qs.values_list(
            'dt_end_cli_work_week', flat=True
        ).distinct().order_by('dt_end_cli_work_week')
        return list(weeks)

    def _get_offices(self):
        """Get list of unique offices in the data."""
        offices = self.qs.values_list(
            'xlc_operation', flat=True
        ).distinct().order_by('xlc_operation')
        return [o for o in offices if o]  # Filter out empty strings

    def _format_week_column(self, date):
        """Format week date as column header (e.g., '9-Nov')."""
        if date is None:
            return ''
        return date.strftime('%#d-%b')  # Windows format

    def _calculate_enrollment_status(self, provisional_count):
        """Determine if employee needs fingerprint enrollment."""
        return "Yes" if provisional_count > self.ENROLLMENT_THRESHOLD else ""

    def _apply_formatting(self, worksheet):
        """Apply BSTT-rpt.xlsx styling to worksheet."""
        if worksheet.max_row == 0:
            return

        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze first row
        worksheet.freeze_panes = 'A2'

    def _queryset_to_dataframe(self, qs, fields):
        """Convert a queryset to a pandas DataFrame."""
        data = list(qs.values(*fields))
        if data:
            return pd.DataFrame(data)
        return pd.DataFrame(columns=fields)

    # ============================================
    # Sheet Writers
    # ============================================

    def _write_checking_file(self, writer):
        """Write metadata/validation sheet."""
        # Get data statistics
        total_records = self.qs.count()
        date_range = self.qs.aggregate(
            min_date=Min('dt_end_cli_work_week'),
            max_date=Max('dt_end_cli_work_week')
        )
        unique_employees = self.qs.values('applicant_id').distinct().count()
        unique_offices = len(self._get_offices())

        # Build metadata
        metadata = [
            ['BSTT Compliance Report - File Information', ''],
            ['', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Year:', self.year],
            ['', ''],
            ['Data Range:', ''],
            ['Start Date:', str(date_range['min_date']) if date_range['min_date'] else 'N/A'],
            ['End Date:', str(date_range['max_date']) if date_range['max_date'] else 'N/A'],
            ['', ''],
            ['Record Counts:', ''],
            ['Total Records:', total_records],
            ['Unique Employees:', unique_employees],
            ['Unique Offices:', unique_offices],
            ['Unique Weeks:', len(self._get_unique_weeks())],
            ['', ''],
            ['Filters Applied:', ''],
        ]

        # Add filter information
        for key, value in self.filters.items():
            if value:
                metadata.append([f'  {key}:', str(value)])

        if not any(v for v in self.filters.values()):
            metadata.append(['  None', ''])

        df = pd.DataFrame(metadata, columns=['Field', 'Value'])
        df.to_excel(writer, sheet_name='CheckingTheFile', index=False)

    def _write_directions(self, writer):
        """Write instructions sheet."""
        directions = [
            ['BSTT Compliance Report - Directions', ''],
            ['', ''],
            ['Sheet Descriptions:', ''],
            ['', ''],
            ['CheckingTheFile', 'Report metadata and filter information'],
            ['Directions', 'This sheet - instructions for using the report'],
            ['Data', 'Consolidated time entry data with aggregations'],
            ['All', 'Entry type pivot by employee - shows count of each entry type per employee'],
            ['Prov', 'Weekly provisional tracking with enrollment status'],
            ['Write Ins', 'Write-in entries tracking by employee'],
            ['Provisional', 'Detailed list of employees with provisional entries'],
            ['[Office Name]', 'Office-specific entry type analysis'],
            ['', ''],
            ['Entry Types:', ''],
            ['Finger', 'Biometric (fingerprint) clock entry - Target: 95%+'],
            ['Provisional Entry', 'Temporary entry pending fingerprint enrollment - Target: <1%'],
            ['Write-In', 'Manual write-in entry - Target: <3%'],
            ['Missing c/o', 'Missing clock-out entry - Target: <2%'],
            ['', ''],
            ['Enrollment of Fingerprint Needed:', ''],
            ['Yes', f'Employee has more than {self.ENROLLMENT_THRESHOLD} provisional entries and needs fingerprint enrollment'],
            ['(blank)', 'Employee does not require immediate enrollment action'],
            ['', ''],
            ['Color Coding:', ''],
            ['Green (95%+)', 'Meeting finger rate target'],
            ['Yellow (90-95%)', 'Below target but acceptable'],
            ['Red (<90%)', 'Requires immediate attention'],
        ]

        df = pd.DataFrame(directions, columns=['Topic', 'Description'])
        df.to_excel(writer, sheet_name='Directions', index=False)

    def _write_data(self, writer):
        """Write consolidated data sheet matching BSTT-rpt.xlsx Data sheet format."""
        # Aggregate data by employee, week, office
        data = self.qs.values(
            'xlc_operation', 'dt_end_cli_work_week', 'applicant_id',
            'shift_number', 'bu_dept_name', 'entry_type', 'full_name'
        ).annotate(
            total_hours=Sum('total_hours'),
            total_entries=Count('id')
        ).order_by('xlc_operation', 'dt_end_cli_work_week', 'full_name')

        if data:
            df = pd.DataFrame(list(data))

            # Rename columns to match BSTT-rpt.xlsx
            df = df.rename(columns={
                'xlc_operation': 'OfcName',
                'dt_end_cli_work_week': 'dtEndCliWorkWeek',
                'applicant_id': 'ApplicantID',
                'shift_number': 'ShiftNumber',
                'bu_dept_name': 'BUDeptName',
                'entry_type': 'EntryType',
                'full_name': 'FullName',
                'total_hours': 'Total Hours',
                'total_entries': 'TotalEntries'
            })

            # Add weekday column (derived from week ending date)
            df['weekday'] = pd.to_datetime(df['dtEndCliWorkWeek']).dt.day_name()

            # Add w-e column (formatted week ending)
            df['w-e'] = pd.to_datetime(df['dtEndCliWorkWeek']).dt.strftime('%#d-%b')

            # Reorder columns
            column_order = [
                'OfcName', 'dtEndCliWorkWeek', 'ApplicantID', 'ShiftNumber',
                'BUDeptName', 'EntryType', 'FullName', 'Total Hours',
                'TotalEntries', 'weekday', 'w-e'
            ]
            df = df[[c for c in column_order if c in df.columns]]

            df.to_excel(writer, sheet_name='Data', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='Data', index=False
            )

    def _write_all_pivot(self, writer):
        """Write entry type pivot sheet - employees x entry types."""
        # Get entry counts per employee per entry type
        data = self.qs.values(
            'full_name', 'entry_type'
        ).annotate(
            count=Count('id')
        ).order_by('full_name')

        if data:
            df = pd.DataFrame(list(data))

            # Pivot the data
            pivot = df.pivot_table(
                index='full_name',
                columns='entry_type',
                values='count',
                aggfunc='sum',
                fill_value=0
            ).reset_index()

            # Ensure standard entry type columns exist
            entry_types = ['Finger', 'Missing c/o', 'Provisional Entry', 'Write-In']
            for et in entry_types:
                if et not in pivot.columns:
                    pivot[et] = 0

            # Calculate grand total
            pivot['Grand Total'] = pivot[entry_types].sum(axis=1)

            # Rename and reorder columns
            pivot = pivot.rename(columns={'full_name': 'Employee'})

            # Sort by Grand Total descending
            pivot = pivot.sort_values('Grand Total', ascending=False)

            # Reorder columns
            column_order = ['Employee'] + entry_types + ['Grand Total']
            pivot = pivot[[c for c in column_order if c in pivot.columns]]

            pivot.to_excel(writer, sheet_name='All', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='All', index=False
            )

    def _write_prov_weekly(self, writer):
        """Write weekly provisional tracking sheet with enrollment status."""
        # Filter to provisional entries only
        prov_qs = self.qs.filter(entry_type='Provisional Entry')

        if not prov_qs.exists():
            pd.DataFrame({'No Data': ['No provisional entries found']}).to_excel(
                writer, sheet_name='Prov', index=False
            )
            return

        # Get unique weeks
        weeks = self._get_unique_weeks()

        # Get counts per employee per office per week
        data = prov_qs.values(
            'xlc_operation', 'full_name', 'dt_end_cli_work_week'
        ).annotate(
            count=Count('id')
        ).order_by('xlc_operation', 'full_name')

        df = pd.DataFrame(list(data))

        # Pivot by week
        pivot = df.pivot_table(
            index=['xlc_operation', 'full_name'],
            columns='dt_end_cli_work_week',
            values='count',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # Rename week columns to formatted dates
        week_cols = [c for c in pivot.columns if c not in ['xlc_operation', 'full_name']]
        rename_map = {w: self._format_week_column(w) for w in week_cols}
        pivot = pivot.rename(columns=rename_map)

        # Calculate totals and average
        formatted_week_cols = [self._format_week_column(w) for w in weeks if w in week_cols]
        if formatted_week_cols:
            pivot['Total'] = pivot[formatted_week_cols].sum(axis=1)
            pivot['Average'] = pivot[formatted_week_cols].mean(axis=1).round(1)
        else:
            pivot['Total'] = 0
            pivot['Average'] = 0

        # Add enrollment status
        pivot['Enrollment of Fingerprint Needed'] = pivot['Total'].apply(
            self._calculate_enrollment_status
        )

        # Rename columns
        pivot = pivot.rename(columns={
            'xlc_operation': 'OfcName',
            'full_name': 'FullName'
        })

        # Reorder columns
        base_cols = ['OfcName', 'FullName']
        end_cols = ['Total', 'Average', 'Enrollment of Fingerprint Needed']
        week_col_order = [self._format_week_column(w) for w in sorted(weeks) if self._format_week_column(w) in pivot.columns]
        column_order = base_cols + week_col_order + end_cols
        pivot = pivot[[c for c in column_order if c in pivot.columns]]

        # Sort by Total descending
        pivot = pivot.sort_values('Total', ascending=False)

        pivot.to_excel(writer, sheet_name='Prov', index=False)

    def _write_write_ins(self, writer):
        """Write write-in entries tracking sheet."""
        # Filter to write-in entries only
        writein_qs = self.qs.filter(entry_type='Write-In')

        if not writein_qs.exists():
            pd.DataFrame({'No Data': ['No write-in entries found']}).to_excel(
                writer, sheet_name='Write Ins', index=False
            )
            return

        # Get unique weeks
        weeks = self._get_unique_weeks()

        # Get counts per employee per office per week
        data = writein_qs.values(
            'xlc_operation', 'full_name', 'dt_end_cli_work_week'
        ).annotate(
            count=Count('id')
        ).order_by('xlc_operation', 'full_name')

        df = pd.DataFrame(list(data))

        # Pivot by week
        pivot = df.pivot_table(
            index=['xlc_operation', 'full_name'],
            columns='dt_end_cli_work_week',
            values='count',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # Rename week columns
        week_cols = [c for c in pivot.columns if c not in ['xlc_operation', 'full_name']]
        rename_map = {w: self._format_week_column(w) for w in week_cols}
        pivot = pivot.rename(columns=rename_map)

        # Calculate totals
        formatted_week_cols = [self._format_week_column(w) for w in weeks if w in week_cols]
        if formatted_week_cols:
            pivot['Total'] = pivot[formatted_week_cols].sum(axis=1)
        else:
            pivot['Total'] = 0

        # Rename columns
        pivot = pivot.rename(columns={
            'xlc_operation': 'OfcName',
            'full_name': 'FullName'
        })

        # Sort by Total descending
        pivot = pivot.sort_values('Total', ascending=False)

        pivot.to_excel(writer, sheet_name='Write Ins', index=False)

    def _write_provisional_detail(self, writer):
        """Write detailed provisional entries list."""
        # Filter to provisional entries only
        prov_qs = self.qs.filter(entry_type='Provisional Entry')

        if not prov_qs.exists():
            pd.DataFrame({'No Data': ['No provisional entries found']}).to_excel(
                writer, sheet_name='Provisional', index=False
            )
            return

        # Get summary per employee
        data = prov_qs.values(
            'xlc_operation', 'full_name', 'applicant_id', 'bu_dept_name'
        ).annotate(
            total_entries=Count('id'),
            total_hours=Sum('total_hours')
        ).order_by('-total_entries')

        df = pd.DataFrame(list(data))

        # Add enrollment status
        df['Enrollment of Fingerprint Needed'] = df['total_entries'].apply(
            self._calculate_enrollment_status
        )

        # Rename columns
        df = df.rename(columns={
            'xlc_operation': 'Office',
            'full_name': 'Employee',
            'applicant_id': 'Employee ID',
            'bu_dept_name': 'Department',
            'total_entries': 'Provisional Count',
            'total_hours': 'Total Hours'
        })

        # Reorder columns
        column_order = ['Office', 'Employee', 'Employee ID', 'Department',
                       'Provisional Count', 'Total Hours', 'Enrollment of Fingerprint Needed']
        df = df[[c for c in column_order if c in df.columns]]

        df.to_excel(writer, sheet_name='Provisional', index=False)

    def _write_office_sheet(self, writer, office_name):
        """Write office-specific entry type analysis sheet."""
        # Filter to specific office
        office_qs = self.qs.filter(xlc_operation=office_name)

        if not office_qs.exists():
            return  # Skip empty offices

        # Get entry counts per employee per entry type
        data = office_qs.values(
            'full_name', 'entry_type'
        ).annotate(
            count=Count('id')
        ).order_by('full_name')

        if not data:
            return

        df = pd.DataFrame(list(data))

        # Pivot the data
        pivot = df.pivot_table(
            index='full_name',
            columns='entry_type',
            values='count',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # Ensure standard entry type columns exist (excluding Finger for this view)
        non_finger_types = ['Missing c/o', 'Provisional Entry', 'Write-In']
        for et in non_finger_types:
            if et not in pivot.columns:
                pivot[et] = 0

        # Calculate grand total (of non-compliance entries)
        existing_cols = [c for c in non_finger_types if c in pivot.columns]
        pivot['Grand Total'] = pivot[existing_cols].sum(axis=1)

        # Add enrollment status based on provisional entries
        if 'Provisional Entry' in pivot.columns:
            pivot['Enrollment of Fingerprint Needed'] = pivot['Provisional Entry'].apply(
                self._calculate_enrollment_status
            )
        else:
            pivot['Enrollment of Fingerprint Needed'] = ''

        # Rename column
        pivot = pivot.rename(columns={'full_name': 'Full Name'})

        # Reorder columns
        column_order = ['Full Name'] + non_finger_types + ['Grand Total', 'Enrollment of Fingerprint Needed']
        pivot = pivot[[c for c in column_order if c in pivot.columns]]

        # Sort by Grand Total descending
        pivot = pivot.sort_values('Grand Total', ascending=False)

        # Only include employees with non-compliance entries
        pivot = pivot[pivot['Grand Total'] > 0]

        if not pivot.empty:
            # Truncate sheet name to 31 chars (Excel limit)
            sheet_name = office_name[:31] if len(office_name) > 31 else office_name
            # Remove invalid characters from sheet name
            invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
            for char in invalid_chars:
                sheet_name = sheet_name.replace(char, '-')
            pivot.to_excel(writer, sheet_name=sheet_name, index=False)

    # ============================================
    # Legacy Methods (for backward compatibility)
    # ============================================

    def _write_summary(self, writer):
        """Write executive summary sheet (legacy)."""
        calc = KPICalculator(self.qs)
        kpis = calc.calculate_all()

        summary_data = [
            ['BSTT Compliance Dashboard - Executive Summary', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['Year:', self.year],
            ['', ''],
            ['COMPLIANCE METRICS', ''],
            ['Finger Rate (Target: 95%)', f"{kpis['finger_rate']}%"],
            ['Provisional Rate (Target: <1%)', f"{kpis['provisional_rate']}%"],
            ['Write-In Rate (Target: <3%)', f"{kpis['write_in_rate']}%"],
            ['Missing C/O Rate (Target: <2%)', f"{kpis['missing_co_rate']}%"],
            ['Manual Entry Rate', f"{kpis['manual_rate']}%"],
            ['', ''],
            ['VOLUME METRICS', ''],
            ['Total Entries', kpis['total_entries']],
            ['Total Hours', kpis['total_hours']],
            ['Unique Employees', kpis['unique_employees']],
            ['Unique Offices', kpis['unique_offices']],
            ['Weeks Covered', kpis['unique_weeks']],
            ['', ''],
            ['EFFICIENCY METRICS', ''],
            ['First-Try Clock-In Rate', f"{kpis['first_try_clock_in_rate']}%"],
            ['First-Try Clock-Out Rate', f"{kpis['first_try_clock_out_rate']}%"],
            ['Avg Clock-In Tries', kpis['avg_clock_in_tries']],
            ['Avg Clock-Out Tries', kpis['avg_clock_out_tries']],
        ]

        df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Summary', index=False)

    def _write_by_office(self, writer):
        """Write office breakdown sheet (legacy)."""
        calc = KPICalculator(self.qs)
        office_data = calc.by_office()

        if office_data:
            df = pd.DataFrame(office_data)
            cols = ['office', 'total_entries', 'finger_rate', 'provisional_rate',
                   'write_in_rate', 'missing_co_rate', 'total_hours', 'unique_employees']
            df = df[[c for c in cols if c in df.columns]]
            df.columns = ['Office', 'Entries', 'Finger %', 'Provisional %',
                         'Write-In %', 'Missing C/O %', 'Total Hours', 'Employees']
            df.to_excel(writer, sheet_name='By Office', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='By Office', index=False
            )

    def _write_by_entry_type(self, writer):
        """Write entry type breakdown sheet (legacy)."""
        entry_types = self.qs.values('entry_type').annotate(
            count=Count('id'),
            total_hours=Sum('total_hours'),
            employees=Count('applicant_id', distinct=True)
        ).order_by('-count')

        if entry_types:
            df = pd.DataFrame(list(entry_types))
            total = df['count'].sum()
            df['percentage'] = (df['count'] / total * 100).round(1)
            df = df[['entry_type', 'count', 'percentage', 'total_hours', 'employees']]
            df.columns = ['Entry Type', 'Count', 'Percentage', 'Total Hours', 'Employees']
            df.to_excel(writer, sheet_name='By Entry Type', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='By Entry Type', index=False
            )

    def _write_weekly_trends(self, writer):
        """Write weekly trends sheet (legacy)."""
        calc = KPICalculator(self.qs)
        weekly_data = calc.by_week()

        if weekly_data:
            df = pd.DataFrame(weekly_data)
            cols = ['week', 'total_entries', 'finger_rate', 'provisional_rate',
                   'total_hours', 'unique_employees']
            df = df[[c for c in cols if c in df.columns]]
            df.columns = ['Week', 'Entries', 'Finger %', 'Provisional %',
                         'Total Hours', 'Employees']
            df.to_excel(writer, sheet_name='Weekly Trends', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='Weekly Trends', index=False
            )

    def _write_employees(self, writer):
        """Write employee summary sheet (legacy)."""
        employees = self.qs.values(
            'applicant_id', 'full_name', 'xlc_operation'
        ).annotate(
            entries=Count('id'),
            total_hours=Sum('total_hours'),
            finger_entries=Count('id', filter=Q(entry_type='Finger'))
        ).order_by('-entries')[:500]

        if employees:
            df = pd.DataFrame(list(employees))
            df['finger_rate'] = (df['finger_entries'] / df['entries'] * 100).round(1)
            df = df[['full_name', 'xlc_operation', 'entries', 'finger_rate', 'total_hours']]
            df.columns = ['Employee', 'Office', 'Entries', 'Finger %', 'Total Hours']
            df.to_excel(writer, sheet_name='Employees', index=False)
        else:
            pd.DataFrame({'No Data': ['No records found']}).to_excel(
                writer, sheet_name='Employees', index=False
            )

    def generate_weekly_summary(self) -> BytesIO:
        """Generate weekly summary report."""
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self._write_summary(writer)
            self._write_weekly_trends(writer)

        output.seek(0)
        return output
